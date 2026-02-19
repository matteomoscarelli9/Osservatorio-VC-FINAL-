#!/usr/bin/env python3
import argparse
import html
import json
import os
import re
import sqlite3
import subprocess
import sys
import difflib
import unicodedata
from datetime import datetime
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import List, Dict, Tuple
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET

try:
    from openpyxl import load_workbook
except Exception as e:
    load_workbook = None

try:
    from openai import OpenAI
except Exception:
    OpenAI = None
try:
    import psycopg
except Exception:
    psycopg = None

SEP = "<<<SEP>>>"
DEFAULT_HQ_CACHE = "/Users/matteomoscarelli/Documents/New project/automations/hq_cache.json"
EXCEL_PATH_DEFAULT = "/Users/matteomoscarelli/Library/CloudStorage/OneDrive-Raccoltecondivise-UnitedVentures/United Ventures - United Ventures/04. Portfolio/Portfolio Team/100. OLD/07. Dealflow Meeting MM/Osservatorio VC Italy _.xlsx"
SHEET_DEFAULT = "Funding rounds (3)"
DB_PATH_DEFAULT = str(Path(__file__).resolve().parents[1] / "db" / "rounds.db")
HQ_ENRICH_MODEL_DEFAULT = os.environ.get("OPENAI_HQ_MODEL", "gpt-4.1-mini")

CITY_ALIAS_TO_EN = {
    "milano": "Milan",
    "milan": "Milan",
    "torino": "Turin",
    "turin": "Turin",
    "roma": "Rome",
    "rome": "Rome",
    "napoli": "Naples",
    "naples": "Naples",
    "firenze": "Florence",
    "florence": "Florence",
    "venezia": "Venice",
    "venice": "Venice",
    "genova": "Genoa",
    "genoa": "Genoa",
    "padova": "Padua",
    "padua": "Padua",
    "bologna": "Bologna",
    "bergamo": "Bergamo",
    "parma": "Parma",
    "pisa": "Pisa",
    "modena": "Modena",
    "trento": "Trento",
    "trieste": "Trieste",
    "brescia": "Brescia",
    "verona": "Verona",
    "vicenza": "Vicenza",
    "poggibonsi": "Poggibonsi",
    "bovisio": "Bovisio",
}

ALLOWED_SECTORS = [
    "Agritech",
    "Biotech",
    "Blue Economy",
    "Cleantech",
    "Climate Tech",
    "Consulting",
    "Consumer Services",
    "Crypto",
    "Cybersecurity",
    "Deep Tech",
    "DevOps",
    "eCommerce",
    "Edtech",
    "Energy",
    "Enterprise Tech",
    "Entraitment",
    "Fashion",
    "Femtech",
    "Fintech",
    "Food",
    "Gaming",
    "HR Tech",
    "Industrial Tech",
    "Insurtech",
    "Life Sciences",
    "Logistics",
    "Media & Ad",
    "Mobility",
    "Quantum",
    "Social Network",
    "Silver Economy",
    "Travel & Hospitality",
    "Web3",
]


def log_info(message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {message}", flush=True)


def _normalize_city_key(value: str) -> str:
    s = str(value or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z\s-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_city_name(value: str) -> str:
    s = str(value or "").strip()
    if not s:
        return ""
    if is_generic_hq(s):
        return ""
    key = _normalize_city_key(s)
    if not key:
        return ""
    if key in CITY_ALIAS_TO_EN:
        return CITY_ALIAS_TO_EN[key]
    # Fuzzy recovery for common typos (e.g., Milnao -> Milano/Milan).
    close = difflib.get_close_matches(key, CITY_ALIAS_TO_EN.keys(), n=1, cutoff=0.75)
    if close:
        return CITY_ALIAS_TO_EN[close[0]]
    # Fallback: title case cleaned token.
    return re.sub(r"\s+", " ", s).strip(" .,-").title()


def run_osascript(script: str) -> str:
    proc = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or "osascript failed")
    return proc.stdout.strip()


def fetch_latest_outlook_message(sender: str, subject_contains: str, recent_days: int) -> Tuple[str, str, str]:
    # Returns (subject, time_received_str, body_plain)
    applescript = f'''
    set subjectKeyword to "{subject_contains}"
    set senderKeyword to "{sender}"
    set cutoffDate to (current date) - ({recent_days} * days)
    tell application "Microsoft Outlook"
        set totalCount to count of messages of inbox
        set scanCount to totalCount
        if scanCount > 800 then set scanCount to 800
        set latestMsg to missing value
        set latestDate to date "01/01/1970"
        repeat with i from 1 to scanCount
            set m to item i of messages of inbox
            try
                set d to time received of m
            on error
                set d to date "01/01/1970"
            end try
            if d < cutoffDate then
                -- skip old
            else
                try
                    set subj to subject of m as string
                on error
                    set subj to ""
                end try
                try
                    set senderStr to sender of m as string
                on error
                    set senderStr to ""
                end try
                if subjectKeyword is "" then
                    set matchesSubject to true
                else
                    ignoring case
                        set matchesSubject to (subj contains subjectKeyword)
                    end ignoring
                end if
                if senderKeyword is "" then
                    set matchesSender to true
                else
                    ignoring case
                        set matchesSender to (senderStr contains senderKeyword)
                    end ignoring
                end if
                if matchesSubject and matchesSender then
                    if d > latestDate then
                        set latestDate to d
                        set latestMsg to m
                    end if
                end if
            end if
        end repeat
        if latestMsg is missing value then return ""
        set outSubject to subject of latestMsg as string
        set outDate to (time received of latestMsg) as string
        set outBody to plain text content of latestMsg
        return outSubject & "{SEP}" & outDate & "{SEP}" & outBody
    end tell
    '''
    out = run_osascript(applescript)
    if not out:
        return "", "", ""
    parts = out.split(SEP)
    if len(parts) < 3:
        raise RuntimeError("Unexpected Outlook output format")
    subject = parts[0].strip()
    time_received = parts[1].strip()
    body = SEP.join(parts[2:]).strip()
    return subject, time_received, body


def list_latest_outlook_messages(limit: int = 50, recent_days: int = 30) -> str:
    applescript = f'''
    set maxItems to {limit}
    set cutoffDate to (current date) - ({recent_days} * days)
    tell application "Microsoft Outlook"
        set targetMessages to messages of inbox
        set out to ""
        set out to ""
        set countSeen to 0
        repeat with m in targetMessages
            try
                set subj to subject of m as string
            on error
                set subj to ""
            end try
            try
                set senderStr to sender of m as string
            on error
                set senderStr to ""
            end try
            try
                set receivedStr to (time received of m) as string
            on error
                set receivedStr to ""
            end try
            try
                set d to time received of m
            on error
                set d to date "01/01/1970"
            end try
            if d ≥ cutoffDate then
                set out to out & subj & " | " & senderStr & " | " & receivedStr & "\\n"
                set countSeen to countSeen + 1
                if countSeen ≥ maxItems then exit repeat
            end if
        end repeat
        return out
    end tell
    '''
    return run_osascript(applescript)


def fetch_current_outlook_message() -> Tuple[str, str, str]:
    # Returns (subject, time_received_str, body_plain) for the currently selected message
    applescript = f'''
    tell application "Microsoft Outlook"
        set sel to selection
        if sel is {{}} then return ""
        set m to item 1 of sel
        set outSubject to subject of m as string
        set outDate to (time received of m) as string
        set outBody to plain text content of m
        return outSubject & "{SEP}" & outDate & "{SEP}" & outBody
    end tell
    '''
    out = run_osascript(applescript)
    if not out:
        return "", "", ""
    parts = out.split(SEP)
    if len(parts) < 3:
        raise RuntimeError("Unexpected Outlook output format")
    subject = parts[0].strip()
    time_received = parts[1].strip()
    body = SEP.join(parts[2:]).strip()
    return subject, time_received, body


def strip_html_to_text(raw: str) -> str:
    if not raw:
        return ""
    txt = raw
    # Keep block structure so section headings and bullets remain separable.
    txt = re.sub(r"(?i)</(h[1-6]|div|section|article|ul|ol|table|tr)>", "\n", txt)
    txt = re.sub(r"(?i)<(h[1-6]|div|section|article|ul|ol|table|tr)[^>]*>", "\n", txt)
    txt = re.sub(r"(?i)<br\\s*/?>", "\n", txt)
    txt = re.sub(r"(?i)</p>", "\n", txt)
    txt = re.sub(r"(?i)</li>", "\n", txt)
    txt = re.sub(r"(?i)<li[^>]*>", "\n• ", txt)
    txt = re.sub(r"<[^>]+>", "", txt)
    txt = html.unescape(txt)
    txt = txt.replace("\xa0", " ")
    txt = re.sub(r"[ \t]+", " ", txt)
    txt = re.sub(r"\n{3,}", "\n\n", txt)
    return txt.strip()


def fetch_latest_rss_message(rss_url: str, subject_contains: str, recent_days: int) -> Tuple[str, str, str]:
    if not rss_url:
        return "", "", ""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/rss+xml, application/xml, text/xml;q=0.9, */*;q=0.8",
        "Cache-Control": "no-cache",
    }
    candidates = []
    base = rss_url.strip()
    candidates.append(base)
    if not base.endswith("/"):
        candidates.append(base + "/")

    # Common feed aliases for Substack/custom domains
    if "/rss" in base:
        root = base.split("/rss", 1)[0]
        candidates.extend([f"{root}/feed", f"{root}/feed/"])
    elif "/feed" in base:
        root = base.split("/feed", 1)[0]
        candidates.extend([f"{root}/rss", f"{root}/rss/"])

    # Keep order but remove duplicates
    seen = set()
    uniq_candidates = []
    for u in candidates:
        if u not in seen:
            uniq_candidates.append(u)
            seen.add(u)

    payload = None
    errors = []
    for url in uniq_candidates:
        try:
            req = Request(url, headers=headers)
            with urlopen(req, timeout=20) as resp:
                payload = resp.read()
            break
        except Exception as e:
            errors.append(f"{url} -> {e}")

    if payload is None:
        raise RuntimeError("RSS fetch failed. Attempts: " + " | ".join(errors))
    root = ET.fromstring(payload)

    now = datetime.now()
    cutoff = now.timestamp() - (recent_days * 86400)
    best = None

    for item in root.findall(".//item"):
        title = (item.findtext("title") or "").strip()
        pub_date = (item.findtext("pubDate") or "").strip()
        description = (item.findtext("description") or "").strip()
        encoded = ""
        for child in list(item):
            if child.tag.lower().endswith("encoded"):
                encoded = (child.text or "").strip()
                break
        body_html = encoded or description
        body_text = strip_html_to_text(body_html)

        ts = 0
        try:
            dt = parsedate_to_datetime(pub_date)
            ts = dt.timestamp()
        except Exception:
            ts = 0

        if subject_contains and subject_contains.lower() not in title.lower():
            continue
        if ts and ts < cutoff:
            continue
        if best is None or ts > best[0]:
            best = (ts, title, pub_date, body_text)

    if best is None:
        return "", "", ""
    return best[1], best[2], best[3]


def normalize_heading(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip()).lower()


def extract_the_money_section(body: str) -> List[str]:
    if not body:
        return []
    # Normalize line endings
    text = body.replace("\r\n", "\n").replace("\r", "\n")
    # Find "The Money" marker.
    # Important: do not consume the entire line because some RSS payloads append the
    # first bullet on the same line as the heading ("The Money 💰• ...").
    money_idx = re.search(r"(?im)^\s*the\s+money\b[^\n•]*", text)
    if not money_idx:
        # Fallback: first "The Money" occurrence that has bullet-like content nearby.
        for m in re.finditer(r"(?i)the\s+money\b", text):
            window = text[m.end() : m.end() + 2000]
            if re.search(r"(?:^|\n)\s*[•\-*]\s*", window) or re.search(
                r"\b(raised|secured|closed|bagged|landed|collected)\b", window, flags=re.IGNORECASE
            ):
                money_idx = m
                break
    if not money_idx:
        return []
    start = money_idx.end()
    tail = text[start:]
    # Defensive normalization: some RSS HTML collapses heading + first bullet on one line
    # (e.g. "The Money 💰• Company..."), which would otherwise drop the first item.
    tail = re.sub(r"\s*•\s*", "\n• ", tail)
    # Stop at next section-like heading (e.g., "The Buzz", "M&A", "Reading this week").
    m = re.search(
        r"\n\s*(?:the\s+[a-z][a-z\s]+|m\s*&\s*a|reading\s+this\s+week)\b[^\n]*\n",
        tail,
        flags=re.IGNORECASE,
    )
    if m:
        tail = tail[: m.start()]

    # Collect bullet lines. Handle bullets that wrap across lines.
    lines = [ln.strip() for ln in tail.split("\n") if ln.strip()]
    bullets = []
    current = ""
    for ln in lines:
        if ln.startswith("•") or ln.startswith("-") or ln.startswith("*"):
            if current:
                bullets.append(current.strip())
            current = re.sub(r"^[•\-*]\s*", "", ln).strip()
        else:
            if current:
                current += " " + ln.strip()
    if current:
        bullets.append(current.strip())

    # Fallback: if no bullets found, treat sentence-like lines as bullets
    if not bullets:
        bullets = lines

    cleaned = []
    for b in bullets:
        # Cut tail sections often merged into the last bullet in RSS plain text.
        b = re.split(r"\bM&A\b|\bReading this week\b", b, flags=re.IGNORECASE)[0].strip()
        if b:
            cleaned.append(b)

    # If a bullet contains multiple "X, ... raised ..." chunks, split by company-starts.
    # Avoid sentence/dot-based splitting because decimal amounts (e.g. €1.7m) include dots.
    split_chunks = []
    company_raised_pattern = re.compile(
        r"(?:^|(?<=[.;!?])\s+|[•\-*]\s*)([A-Z][A-Za-z0-9&'().+\- ]{1,80},[^\n]*?\braised\b)",
        flags=re.IGNORECASE,
    )
    for b in cleaned:
        starts = [m.start(1) for m in company_raised_pattern.finditer(b)]
        if len(starts) <= 1:
            split_chunks.append(b)
            continue

        starts.append(len(b))
        for i in range(len(starts) - 1):
            chunk = b[starts[i]:starts[i + 1]].strip(" \t\n\r•-")
            if chunk:
                split_chunks.append(chunk)

    def has_funding_signal(text: str) -> bool:
        if not re.search(r"\b(raised|secured|closed|bagged|landed|collected)\b", text, flags=re.IGNORECASE):
            return False
        # Accept euro symbol, eur token, or compact amount formats like 7m / 550k.
        if "€" in text or re.search(r"\beur\b", text, flags=re.IGNORECASE):
            return True
        return re.search(r"\b\d+(?:[.,]\d+)?\s*[mk]\b", text, flags=re.IGNORECASE) is not None

    def has_mna_signal(text: str) -> bool:
        return re.search(
            r"\b(acquired|acquires|acquisition|majority\s+stake|minority\s+stake|merger|merged)\b",
            text,
            flags=re.IGNORECASE,
        ) is not None

    # Strictly keep The Money and exclude M&A-like items if they leak from malformed HTML.
    split_chunks = [b for b in split_chunks if not has_mna_signal(b)]

    # Keep funding-like items when available (The Money rounds).
    funding_bullets = [b for b in split_chunks if has_funding_signal(b)]
    return funding_bullets if funding_bullets else split_chunks


def parse_outlook_datetime(date_str: str) -> datetime | None:
    if not date_str:
        return None
    s = date_str.strip()

    it_months = {
        "gennaio": "January",
        "febbraio": "February",
        "marzo": "March",
        "aprile": "April",
        "maggio": "May",
        "giugno": "June",
        "luglio": "July",
        "agosto": "August",
        "settembre": "September",
        "ottobre": "October",
        "novembre": "November",
        "dicembre": "December",
    }
    it_weekdays = [
        "lunedì",
        "lunedi",
        "martedì",
        "martedi",
        "mercoledì",
        "mercoledi",
        "giovedì",
        "giovedi",
        "venerdì",
        "venerdi",
        "sabato",
        "domenica",
    ]

    s_lower = s.lower()
    for w in it_weekdays:
        s_lower = s_lower.replace(w, "").strip()
    s = s_lower
    for it, en in it_months.items():
        s = re.sub(rf"\\b{it}\\b", en, s, flags=re.IGNORECASE)
    s = s.replace("date ", " ")
    s = s.replace("alle ore", " ")
    s = s.replace(",", " ")
    s = re.sub(r"\\s+", " ", s).strip()

    for fmt in [
        "%A %d %B %Y %H:%M:%S",
        "%A %d %B %Y %H:%M",
        "%d %B %Y %H:%M:%S",
        "%d %B %Y %H:%M",
        "%b %d %Y %I:%M:%S %p",
        "%b %d %Y %I:%M %p",
    ]:
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def parse_date_to_month_year(date_str: str) -> str:
    try:
        dt = parsedate_to_datetime(date_str)
        return dt.strftime("%b %Y")
    except Exception:
        pass
    dt = parse_outlook_datetime(date_str)
    if dt:
        return dt.strftime("%b %Y")
    # Fallback formats.
    for fmt in [
        "%A, %d %B %Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d %B %Y %H:%M:%S",
        "%d %B %Y %H:%M",
        "%b %d, %Y %I:%M:%S %p",
    ]:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%b %Y")
        except Exception:
            pass
    return ""


def infer_quarter(month_year: str) -> str:
    try:
        dt = datetime.strptime(month_year, "%b %Y")
        q = (dt.month - 1) // 3 + 1
        return f"Q{q} {dt.year}"
    except Exception:
        return ""


def openai_extract_rows(bullets: List[str], headers: List[str], email_date: str, model: str) -> List[Dict[str, str]]:
    if OpenAI is None:
        raise RuntimeError("openai package not installed")
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY not set")

    client = OpenAI(api_key=api_key)

    month_year = parse_date_to_month_year(email_date)
    quarter = infer_quarter(month_year) if month_year else ""

    system = (
        "You extract structured funding deals from a newsletter. "
        "Return JSON only."
    )

    user = {
        "headers": headers,
        "email_date": email_date,
        "default_date": month_year,
        "default_quarter": quarter,
        "allowed_sectors": ALLOWED_SECTORS,
        "bullets": bullets,
        "rules": [
            "Return a JSON array. Each item corresponds to one bullet/deal.",
            "Keys must exactly match the headers list. Use empty string for missing data.",
            "Map company name to 'Company'.",
            "Map sector or description to 'Sector 1' using ONLY allowed_sectors. Always choose the closest fit; do not leave empty.",
            "Leave 'Tag' empty.",
            "Set 'HQ' to the city if explicitly available in the bullet; otherwise leave it empty.",
            "Map round size in € millions to 'Round size (€M)' as a number string.",
            "Set 'Date' to the month+year of the email if not explicit (e.g., 'Feb 2026').",
            "Set 'Q' from the Date if possible (e.g., 'Q1 2026').",
            "Map lead investor to 'Lead' and other investors to 'Co-lead / follow 1', 'follow 2', 'follow 3', 'follow 4' in order.",
            "Do not set 'Tag' (leave empty).",
        ],
    }

    def _extract(input_user: Dict[str, object]) -> List[Dict[str, str]]:
        resp = client.responses.create(
            model=model,
            input=[
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(input_user, ensure_ascii=False)},
            ],
        )

        text = ""
        if hasattr(resp, "output_text") and resp.output_text:
            text = resp.output_text
        else:
            for item in getattr(resp, "output", []):
                if getattr(item, "type", "") == "message":
                    for c in getattr(item, "content", []):
                        if getattr(c, "type", "") == "output_text":
                            text += c.text

        text = text.strip()
        if not text:
            raise RuntimeError("Empty response from OpenAI")
        parsed = json.loads(text)
        if not isinstance(parsed, list):
            raise RuntimeError("OpenAI output is not a list")
        return parsed

    rows = _extract(user)
    if bullets and len(rows) < len(bullets):
        retry_user = dict(user)
        retry_user["rules"] = list(user["rules"]) + [
            f"You MUST return exactly {len(bullets)} items (one per bullet, preserving order)."
        ]
        rows = _extract(retry_user)
    return rows


def find_header_row(ws, header_name: str) -> int:
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == header_name.lower():
                return cell.row
    raise RuntimeError(f"Header '{header_name}' not found")


def read_headers(ws, header_row: int) -> List[str]:
    headers = []
    for cell in ws[header_row]:
        if cell.value is None:
            headers.append("")
        else:
            headers.append(str(cell.value).strip())
    # Trim trailing empty headers
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def get_last_data_row(ws, header_row: int, company_col_idx: int) -> int:
    last = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=company_col_idx).value
        if val not in (None, ""):
            last = row
    return last


def is_duplicate(ws, header_row: int, company_col_idx: int, date_col_idx: int, company: str, date_val: str) -> bool:
    if not company or not date_val:
        return False
    company_norm = str(company).strip().lower()
    date_norm = str(date_val).strip().lower()

    last = get_last_data_row(ws, header_row, company_col_idx)
    start = max(header_row + 1, last - 5)
    for row in range(start, last + 1):
        c = ws.cell(row=row, column=company_col_idx).value
        d = ws.cell(row=row, column=date_col_idx).value
        if c is None or d is None:
            continue
        if str(c).strip().lower() == company_norm and str(d).strip().lower() == date_norm:
            return True
    return False


def is_duplicate_with_amount(
    ws,
    header_row: int,
    company_col_idx: int,
    date_col_idx: int,
    amount_col_idx: int,
    company: str,
    date_val: str,
    amount_val: str,
) -> bool:
    if not company or not date_val:
        return False
    company_norm = str(company).strip().lower()
    date_norm = str(date_val).strip().lower()
    amount_norm = str(amount_val or "").strip().replace(",", ".")

    last = get_last_data_row(ws, header_row, company_col_idx)
    start = max(header_row + 1, last - 20)
    for row in range(start, last + 1):
        c = ws.cell(row=row, column=company_col_idx).value
        d = ws.cell(row=row, column=date_col_idx).value
        a = ws.cell(row=row, column=amount_col_idx).value
        if c is None or d is None:
            continue
        if str(c).strip().lower() != company_norm:
            continue
        if str(d).strip().lower() != date_norm:
            continue
        if amount_norm:
            if str(a or "").strip().replace(",", ".") == amount_norm:
                return True
        else:
            return True
    return False


def append_rows(
    ws,
    header_row: int,
    headers: List[str],
    rows: List[Dict[str, str]],
    dedup_company: str,
    dedup_date: str,
) -> tuple[int, List[str]]:
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}
    if dedup_company not in header_map or dedup_date not in header_map:
        raise RuntimeError("Dedup headers not found in sheet")

    company_col_idx = header_map[dedup_company]
    date_col_idx = header_map[dedup_date]

    amount_col_idx = header_map.get("Round size (€M)")
    last = get_last_data_row(ws, header_row, company_col_idx)
    inserted = 0
    inserted_companies = []

    for row in rows:
        company = row.get(dedup_company, "")
        date_val = row.get(dedup_date, "")
        amount_val = row.get("Round size (€M)", "")
        if amount_col_idx:
            dup = is_duplicate_with_amount(
                ws,
                header_row,
                company_col_idx,
                date_col_idx,
                amount_col_idx,
                company,
                date_val,
                amount_val,
            )
        else:
            dup = is_duplicate(ws, header_row, company_col_idx, date_col_idx, company, date_val)
        if dup:
            continue
        last += 1
        for h, col in header_map.items():
            if h in row:
                ws.cell(row=last, column=col).value = row[h]
        inserted += 1
        if company:
            inserted_companies.append(str(company))
    return inserted, inserted_companies


def db_read_headers(db_path: str, database_url: str = "") -> List[str]:
    if database_url:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed but --database-url was provided")
        conn = psycopg.connect(database_url)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = 'rounds'
            ORDER BY ordinal_position
            """
        )
        headers = [row[0] for row in cur.fetchall()]
        cur.close()
        conn.close()
    else:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(rounds)")
        headers = [row[1] for row in cur.fetchall()]
        conn.close()
    if not headers:
        raise RuntimeError("Table 'rounds' not found in DB")
    return headers


def db_read_company_hq_map(db_path: str, database_url: str = "") -> Dict[str, str]:
    out: Dict[str, str] = {}
    if database_url:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed but --database-url was provided")
        conn = psycopg.connect(database_url)
    else:
        conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    try:
        cur.execute('SELECT id, "Company", "HQ" FROM rounds WHERE "Company" IS NOT NULL AND "HQ" IS NOT NULL')
        stats: Dict[str, Dict[str, Tuple[int, int]]] = {}
        for rid, company, hq in cur.fetchall():
            c = str(company or "").strip().lower()
            h = normalize_city_name(hq)
            if not c or not h:
                continue
            stats.setdefault(c, {})
            cnt, last_id = stats[c].get(h, (0, -1))
            stats[c][h] = (cnt + 1, max(last_id, int(rid or 0)))
        for c, cities in stats.items():
            chosen = sorted(cities.items(), key=lambda kv: (kv[1][0], kv[1][1]), reverse=True)[0][0]
            out[c] = chosen
    finally:
        cur.close()
        conn.close()
    return out


def db_read_company_sector_map(db_path: str, database_url: str = "") -> Dict[str, str]:
    out: Dict[str, str] = {}
    if database_url:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed but --database-url was provided")
        conn = psycopg.connect(database_url)
        cur = conn.cursor()
        try:
            cur.execute(
                """
                WITH ranked AS (
                  SELECT
                    LOWER("Company") AS company_key,
                    "Sector 1" AS sector,
                    COUNT(*) AS cnt,
                    MAX(id) AS last_id,
                    ROW_NUMBER() OVER (
                      PARTITION BY LOWER("Company")
                      ORDER BY COUNT(*) DESC, MAX(id) DESC
                    ) AS rn
                  FROM rounds
                  WHERE COALESCE("Company", '') <> ''
                    AND COALESCE("Sector 1", '') <> ''
                  GROUP BY LOWER("Company"), "Sector 1"
                )
                SELECT company_key, sector
                FROM ranked
                WHERE rn = 1
                """
            )
            for company_key, sector in cur.fetchall():
                c = str(company_key or "").strip()
                s = str(sector or "").strip()
                if c and s:
                    out[c] = s
        finally:
            cur.close()
            conn.close()
        return out

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            WITH ranked AS (
              SELECT
                LOWER("Company") AS company_key,
                "Sector 1" AS sector,
                COUNT(*) AS cnt,
                MAX(id) AS last_id,
                ROW_NUMBER() OVER (
                  PARTITION BY LOWER("Company")
                  ORDER BY COUNT(*) DESC, MAX(id) DESC
                ) AS rn
              FROM rounds
              WHERE COALESCE("Company", '') <> ''
                AND COALESCE("Sector 1", '') <> ''
              GROUP BY LOWER("Company"), "Sector 1"
            )
            SELECT company_key, sector
            FROM ranked
            WHERE rn = 1
            """
        )
        for company_key, sector in cur.fetchall():
            c = str(company_key or "").strip()
            s = str(sector or "").strip()
            if c and s:
                out[c] = s
    finally:
        cur.close()
        conn.close()
    return out


def db_read_hq_overrides(db_path: str, database_url: str = "") -> Dict[str, str]:
    out: Dict[str, str] = {}
    if database_url:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed but --database-url was provided")
        conn = psycopg.connect(database_url)
        pg_mode = True
    else:
        conn = sqlite3.connect(db_path)
        pg_mode = False
    cur = conn.cursor()
    try:
        if pg_mode:
            cur.execute("SELECT to_regclass('public.hq_overrides')")
            exists = cur.fetchone()[0] is not None
        else:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='hq_overrides'")
            exists = cur.fetchone() is not None
        if not exists:
            return {}
        cur.execute('SELECT "Company", "HQ" FROM hq_overrides WHERE "Company" IS NOT NULL AND "HQ" IS NOT NULL')
        for company, hq in cur.fetchall():
            c = str(company or "").strip().lower()
            h = normalize_city_name(hq)
            if not c or not h:
                continue
            out[c] = h
    finally:
        cur.close()
        conn.close()
    return out


def db_upsert_hq_overrides(db_path: str, database_url: str, overrides: Dict[str, str]) -> None:
    if not overrides:
        return
    if database_url:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed but --database-url was provided")
        conn = psycopg.connect(database_url)
        cur = conn.cursor()
        try:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS public.hq_overrides (
                    "Company" text PRIMARY KEY,
                    "HQ" text NOT NULL
                )
                """
            )
            for company, city in overrides.items():
                city_norm = normalize_city_name(city)
                if not city_norm:
                    continue
                cur.execute(
                    """
                    INSERT INTO public.hq_overrides ("Company", "HQ")
                    VALUES (%s, %s)
                    ON CONFLICT ("Company") DO UPDATE SET "HQ" = EXCLUDED."HQ"
                    """,
                    (company, city_norm),
                )
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS hq_overrides (
                "Company" TEXT PRIMARY KEY,
                "HQ" TEXT NOT NULL
            )
            """
        )
        for company, city in overrides.items():
            city_norm = normalize_city_name(city)
            if not city_norm:
                continue
            cur.execute(
                """
                INSERT INTO hq_overrides ("Company", "HQ")
                VALUES (?, ?)
                ON CONFLICT("Company") DO UPDATE SET "HQ" = excluded."HQ"
                """,
                (company, city_norm),
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def _qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def db_is_duplicate(conn, company: str, date_val: str, pg_mode: bool = False) -> bool:
    return db_is_duplicate_with_amount(conn, company, date_val, "", pg_mode)


def db_is_duplicate_with_amount(conn, company: str, date_val: str, amount_val: str, pg_mode: bool = False) -> bool:
    if not company or not date_val:
        return False
    p = "%s" if pg_mode else "?"
    cur = conn.cursor()
    if str(amount_val or "").strip():
        cur.execute(
            f'SELECT 1 FROM rounds WHERE LOWER("Company") = LOWER({p}) AND LOWER("Date") = LOWER({p}) '
            f'AND REPLACE(COALESCE("Round size (€M)", \'\'), \',\', \'.\') = REPLACE(COALESCE({p}, \'\'), \',\', \'.\') '
            f'LIMIT 1',
            (str(company).strip(), str(date_val).strip(), str(amount_val).strip()),
        )
    else:
        cur.execute(
            f'SELECT 1 FROM rounds WHERE LOWER("Company") = LOWER({p}) AND LOWER("Date") = LOWER({p}) LIMIT 1',
            (str(company).strip(), str(date_val).strip()),
        )
    out = cur.fetchone() is not None
    cur.close()
    return out


def db_insert_rows(
    db_path: str,
    headers: List[str],
    rows: List[Dict[str, str]],
    dedup_company: str,
    dedup_date: str,
    database_url: str = "",
) -> tuple[int, List[str]]:
    if dedup_company not in headers or dedup_date not in headers:
        raise RuntimeError("Dedup headers not found in DB table")

    insertable_headers = [h for h in headers if h != "id"]
    columns_sql = ", ".join(_qident(h) for h in insertable_headers)
    pg_mode = bool(database_url)
    placeholders_sql = ", ".join(["%s" if pg_mode else "?"] * len(insertable_headers))
    insert_sql = f"INSERT INTO rounds ({columns_sql}) VALUES ({placeholders_sql})"

    if pg_mode:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed but --database-url was provided")
        conn = psycopg.connect(database_url)
    else:
        conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    inserted = 0
    inserted_companies = []
    try:
        for row in rows:
            company = row.get(dedup_company, "")
            date_val = row.get(dedup_date, "")
            amount_val = row.get("Round size (€M)", "")
            if db_is_duplicate_with_amount(conn, company, date_val, amount_val, pg_mode):
                continue
            values = [row.get(h, "") for h in insertable_headers]
            cur.execute(insert_sql, values)
            inserted += 1
            if company:
                inserted_companies.append(str(company))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return inserted, inserted_companies


def load_hq_cache(path: str) -> Dict[str, str]:
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {str(k).strip().lower(): str(v).strip() for k, v in data.items()}
    except Exception:
        pass
    return {}


def normalize_sector(value: str) -> str:
    if not value:
        return ""
    val = value.strip().lower()
    # Map common variants to allowed sectors
    variants = {
        "clean tech": "Cleantech",
        "cleantech": "Cleantech",
        "climate tech": "Climate Tech",
        "cyber security": "Cybersecurity",
        "cybersecurity": "Cybersecurity",
        "e-commerce": "eCommerce",
        "ecommerce": "eCommerce",
        "enterprise tech": "Enterprise Tech",
        "industrial tech": "Industrial Tech",
        "life sciences": "Life Sciences",
        "fintech": "Fintech",
        "insurtech": "Insurtech",
        "hr tech": "HR Tech",
        "foodtech": "Food",
        "food tech": "Food",
        "edtech": "Edtech",
        "web 3": "Web3",
        "social network": "Social Network",
    }
    if val in variants:
        return variants[val]
    for s in ALLOWED_SECTORS:
        if s.lower() == val:
            return s
    return ""


def infer_sector_from_bullet(company: str, bullets: List[str]) -> str:
    if not company:
        return ""
    comp = company.strip().lower()
    bullet = ""
    for b in bullets:
        if comp in b.lower():
            bullet = b.lower()
            break
    if not bullet:
        return ""
    keywords = [
        (["nuclear", "energy", "power", "fusion", "reactor"], "Energy"),
        (["fintech", "bank", "payments", "payment", "insurance", "insurtech"], "Fintech"),
        (["caregiver", "caregivers", "health", "medical", "clinic", "hospital", "biotech", "pharma", "medtech"], "Life Sciences"),
        (["ai", "machine learning", "ml", "deep tech", "quantum"], "Deep Tech"),
        (["cyber", "security"], "Cybersecurity"),
        (["ecommerce", "e-commerce", "marketplace", "retail"], "eCommerce"),
        (["logistics", "supply chain", "delivery"], "Logistics"),
        (["mobility", "transport", "fleet", "automotive"], "Mobility"),
        (["climate", "cleantech", "carbon"], "Climate Tech"),
        (["enterprise", "saas", "b2b", "devops"], "Enterprise Tech"),
        (["industrial", "manufacturing", "factory"], "Industrial Tech"),
        (["food", "agri", "agritech"], "Food"),
        (["travel", "hospitality"], "Travel & Hospitality"),
        (["web3", "crypto", "blockchain"], "Web3"),
    ]
    for keys, sector in keywords:
        if any(k in bullet for k in keys):
            return sector
    return ""


def format_decimal_comma(value: str) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    # If already uses comma or is not numeric-like, return as-is
    if "," in s:
        return s
    # Accept numbers like 1.23 or 1 or 1.0
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return s.replace(".", ",")
    return s


def extract_company_from_bullet(bullet: str) -> str:
    if not bullet:
        return ""
    s = bullet.strip()
    s = re.sub(r"^[^A-Za-z0-9]+", "", s)
    m = re.match(r"^\s*([A-Za-z0-9&'().+\- ]+?)[,:\-]\s", s)
    if m:
        name = m.group(1).strip()
        return re.sub(r"\s+", " ", name)
    # Fallback: capture first title-like token sequence before "raised".
    m = re.search(r"([A-Z][A-Za-z0-9&'().+\- ]{1,80}?)\s*,[^.]{0,180}\braised\b", s, flags=re.IGNORECASE)
    if m:
        return re.sub(r"\s+", " ", m.group(1).strip())
    m = re.search(
        r"^\s*([A-Z][A-Za-z0-9&'().+\- ]{1,80}?)\s+(?:raised|secured|closed|bagged|landed|collected)\b",
        s,
        flags=re.IGNORECASE,
    )
    if m:
        return re.sub(r"\s+", " ", m.group(1).strip())
    m = re.match(r"^\s*([A-Z][A-Za-z0-9&'().+\- ]{1,80})\s", s)
    return re.sub(r"\s+", " ", m.group(1).strip()) if m else ""


def extract_amount_from_bullet(bullet: str) -> str:
    if not bullet:
        return ""
    txt = bullet.lower().replace(",", ".")
    m = re.search(r"€\s*([0-9]+(?:\.[0-9]+)?)\s*([mk])\b", txt)
    if m:
        n = float(m.group(1))
        unit = m.group(2)
        if unit == "k":
            n = n / 1000.0
        out = f"{n:.3f}".rstrip("0").rstrip(".")
        return out.replace(".", ",")
    m = re.search(r"€\s*([0-9]+(?:\.[0-9]+)?)", txt)
    if m:
        return m.group(1).replace(".", ",")
    return ""


def extract_investors_from_bullet(bullet: str) -> List[str]:
    if not bullet:
        return []
    m = re.search(r"\bfrom\s+(.+?)(?:\.\s*$|$)", bullet, flags=re.IGNORECASE)
    if not m:
        return []
    tail = m.group(1).strip()
    tail = re.sub(r"\band\b", ",", tail, flags=re.IGNORECASE)
    investors = [x.strip(" .") for x in tail.split(",")]
    return [x for x in investors if x]


def infer_hq_from_bullet(bullet: str) -> str:
    if not bullet:
        return ""
    s = bullet.strip()
    city_phrase = r"(?:[A-Z][A-Za-zÀ-ÖØ-öø-ÿ'.-]{1,30}(?:\s+[A-Z][A-Za-zÀ-ÖØ-öø-ÿ'.-]{1,30}){0,2})"
    patterns = [
        rf"\b({city_phrase})-based\b",
        rf"\bbased in\s+({city_phrase})\b",
        r"\bcon sede a\s+([A-Z][A-Za-zÀ-ÖØ-öø-ÿ' .-]{1,40})\b",
        r"\bcon sede in\s+([A-Z][A-Za-zÀ-ÖØ-öø-ÿ' .-]{1,40})\b",
        rf"\bheadquartered in\s+({city_phrase})\b",
    ]
    for pat in patterns:
        m = re.search(pat, s, flags=re.IGNORECASE)
        if m:
            city = re.sub(r"\s+", " ", m.group(1)).strip(" .,-")
            if city and len(city) >= 2:
                return city
    return ""


def find_company_bullet(company: str, bullets: List[str]) -> str:
    comp = str(company or "").strip().lower()
    if not comp:
        return ""
    for b in bullets:
        if comp in str(b).lower():
            return b
    return ""


def is_generic_hq(value: str) -> bool:
    v = str(value or "").strip().lower()
    return v in ("", "italy", "italia", "<city>", "city", "unknown", "n/a", "na", "nd")


def resolve_hq(
    company: str,
    current_hq: str,
    bullet: str,
    hq_overrides: Dict[str, str],
    hq_cache: Dict[str, str],
    db_hq_map: Dict[str, str],
) -> str:
    cur = str(current_hq or "").strip()
    if cur and not is_generic_hq(cur):
        return normalize_city_name(cur) or cur
    key = str(company or "").strip().lower()
    if key and key in hq_overrides and hq_overrides[key]:
        return normalize_city_name(hq_overrides[key]) or hq_overrides[key]
    if key and key in hq_cache and hq_cache[key]:
        return normalize_city_name(hq_cache[key]) or hq_cache[key]
    if key and key in db_hq_map and db_hq_map[key]:
        return normalize_city_name(db_hq_map[key]) or db_hq_map[key]
    inferred = infer_hq_from_bullet(bullet)
    if inferred:
        return normalize_city_name(inferred) or inferred
    return "Italy"


def openai_enrich_hq_overrides(
    rows: List[Dict[str, str]],
    bullets: List[str],
    model: str = HQ_ENRICH_MODEL_DEFAULT,
) -> Dict[str, str]:
    if OpenAI is None:
        return {}
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {}

    unresolved = []
    seen = set()
    for r in rows:
        company = str(r.get("Company", "")).strip()
        if not company:
            continue
        if not is_generic_hq(r.get("HQ", "")):
            continue
        key = company.lower()
        if key in seen:
            continue
        seen.add(key)
        unresolved.append({"company": company, "bullet": find_company_bullet(company, bullets)})
    if not unresolved:
        return {}

    client = OpenAI(api_key=api_key)
    system = (
        "Find the headquarters city for each company. "
        "Use web sources when needed (company website, LinkedIn, trusted profiles). "
        "Return JSON array only, one item per input item, preserving order. "
        'Each item keys: "company", "city". '
        "City must be a concrete city name only (no country, no placeholders). "
        "If uncertain, return empty city."
    )
    user = {"items": unresolved}

    try:
        request_payload = {
            "model": model,
            "input": [
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(user, ensure_ascii=False)},
            ],
        }
        # Try with web search enabled first.
        try:
            resp = client.responses.create(
                **request_payload,
                tools=[{"type": "web_search_preview"}],
            )
        except Exception:
            resp = client.responses.create(**request_payload)
        text = resp.output_text if hasattr(resp, "output_text") and resp.output_text else ""
        if not text:
            log_info("HQ enrichment: empty response from OpenAI")
            return {}
        parsed = json.loads(text)
        if not isinstance(parsed, list):
            log_info("HQ enrichment: invalid non-list JSON response")
            return {}
        out = {}
        for item in parsed:
            if not isinstance(item, dict):
                continue
            company = str(item.get("company", "")).strip()
            city = re.sub(r"\s+", " ", str(item.get("city", "")).strip()).strip(" .,-")
            city = normalize_city_name(city)
            if not company or not city or is_generic_hq(city):
                continue
            out[company] = city
        return out
    except Exception as e:
        log_info(f"HQ enrichment failed: {e}")
        return {}


def synthesize_rows_for_missing_companies(
    bullets: List[str],
    rows: List[Dict[str, str]],
    headers: List[str],
    date_value: str,
    hq_overrides: Dict[str, str] | None = None,
    hq_cache: Dict[str, str] | None = None,
    db_hq_map: Dict[str, str] | None = None,
) -> List[Dict[str, str]]:
    hq_overrides = hq_overrides or {}
    hq_cache = hq_cache or {}
    db_hq_map = db_hq_map or {}
    existing = {str(r.get("Company", "")).strip().lower() for r in rows if str(r.get("Company", "")).strip()}
    synthesized = []
    for b in bullets:
        if not re.search(r"\b(raised|secured|closed|bagged|landed|collected)\b", b, flags=re.IGNORECASE):
            continue
        company = extract_company_from_bullet(b)
        if not company:
            continue
        if company.strip().lower() in existing:
            continue

        row = {h: "" for h in headers if h}
        row["Company"] = company
        row["Date"] = date_value
        row["Q"] = infer_quarter(date_value)
        row["Round size (€M)"] = extract_amount_from_bullet(b)
        if not row["Round size (€M)"]:
            continue
        row["Sector 1"] = infer_sector_from_bullet(company, [b])
        row["HQ"] = resolve_hq(company, "", b, hq_overrides, hq_cache, db_hq_map)

        inv = extract_investors_from_bullet(b)
        if inv:
            row["Lead"] = inv[0]
        if len(inv) > 1:
            row["Co-lead / follow 1"] = inv[1]
        if len(inv) > 2:
            row["follow 2"] = inv[2]
        if len(inv) > 3:
            row["follow 3"] = inv[3]
        if len(inv) > 4:
            row["follow 4"] = inv[4]

        synthesized.append(row)
        existing.add(company.strip().lower())
    return synthesized


def backfill_hq_from_cache(ws, header_row: int, headers: List[str], hq_cache: Dict[str, str]) -> int:
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}
    if "Company" not in header_map or "HQ" not in header_map:
        return 0
    company_col = header_map["Company"]
    hq_col = header_map["HQ"]
    updated = 0
    for row in range(header_row + 1, ws.max_row + 1):
        company = ws.cell(row=row, column=company_col).value
        hq = ws.cell(row=row, column=hq_col).value
        if not company:
            continue
        comp_key = str(company).strip().lower()
        if str(hq).strip().lower() == "italy" and comp_key in hq_cache:
            ws.cell(row=row, column=hq_col).value = hq_cache[comp_key]
            updated += 1
    return updated


def main():
    parser = argparse.ArgumentParser(description="Parse DealflowIT newsletter and append to Excel.")
    parser.add_argument("--path", default=EXCEL_PATH_DEFAULT, help="Path to Excel file")
    parser.add_argument("--sheet", default=SHEET_DEFAULT, help="Sheet name")
    parser.add_argument("--db", default="", help=f"SQLite DB path for direct insert (default: {DB_PATH_DEFAULT})")
    parser.add_argument("--database-url", default="", help="Postgres connection URL for direct insert")
    parser.add_argument("--rss-url", default="", help="RSS URL to read the latest newsletter from")
    parser.add_argument("--sender", default="")
    parser.add_argument("--subject", default="TWIS")
    parser.add_argument("--model", default="gpt-5.2")
    parser.add_argument("--hq-model", default=HQ_ENRICH_MODEL_DEFAULT, help="OpenAI model for HQ city enrichment")
    parser.add_argument("--after", default="", help="Filter emails received on/after this date (YYYY-MM-DD)")
    parser.add_argument("--before", default="", help="Filter emails received on/before this date (YYYY-MM-DD)")
    parser.add_argument("--recent-days", type=int, default=30, help="Only scan emails from the last N days")
    parser.add_argument("--hq-cache", default=DEFAULT_HQ_CACHE, help="Path to HQ cache JSON")
    parser.add_argument("--use-current", action="store_true", help="Use current selected Outlook message")
    parser.add_argument("--list-latest", type=int, default=0, help="List latest N messages (subject | sender) and exit")
    parser.add_argument("--backfill-hq", action="store_true", help="Replace HQ='Italy' using HQ cache in existing rows")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()
    db_mode = bool(args.db.strip() or args.database_url.strip())
    db_path = args.db.strip() or DB_PATH_DEFAULT
    database_url = args.database_url.strip()
    log_info(
        f"Start run: sheet='{args.sheet}', recent_days={args.recent_days}, "
        f"use_current={args.use_current}, dry_run={args.dry_run}, backfill_hq={args.backfill_hq}"
    )
    if db_mode:
        if database_url:
            log_info("DB mode enabled. Target: PostgreSQL")
        else:
            log_info(f"DB mode enabled. DB path: {db_path}")
    else:
        log_info(f"Excel path: {args.path}")

    if not db_mode and load_workbook is None:
        raise RuntimeError("openpyxl not installed")

    if args.list_latest and args.list_latest > 0:
        log_info(f"Listing latest {args.list_latest} Outlook messages")
        print(list_latest_outlook_messages(args.list_latest, args.recent_days))
        return

    if args.rss_url:
        log_info(f"Reading latest newsletter from RSS: {args.rss_url}")
        subject, time_received, body = fetch_latest_rss_message(args.rss_url, args.subject, args.recent_days)
    elif args.use_current:
        log_info("Reading currently selected Outlook message")
        subject, time_received, body = fetch_current_outlook_message()
    else:
        log_info(f"Searching latest Outlook message (subject contains '{args.subject}', sender contains '{args.sender}')")
        subject, time_received, body = fetch_latest_outlook_message(args.sender, args.subject, args.recent_days)
    if not subject:
        log_info("No matching message found")
        print("No matching email found")
        return
    log_info(f"Message selected: subject='{subject}' time_received='{time_received}'")
    if args.debug:
        print(f"DEBUG subject: {subject}")
        print(f"DEBUG time_received: {time_received}")
        print("DEBUG body head:")
        print(body[:1000])

    if args.after or args.before:
        dt = parse_outlook_datetime(time_received)
        if dt is None:
            print("Could not parse Outlook date; skipping date filter")
        else:
            after_dt = datetime.strptime(args.after, "%Y-%m-%d") if args.after else None
            before_dt = datetime.strptime(args.before, "%Y-%m-%d") if args.before else None
            if after_dt and dt.date() < after_dt.date():
                print("No matching email found in date range")
                return
            if before_dt and dt.date() > before_dt.date():
                print("No matching email found in date range")
                return

    bullets = extract_the_money_section(body)
    if not bullets:
        log_info("No 'The Money' section extracted from message body")
        print("No 'The Money' items found")
        if args.debug:
            print("DEBUG body head (no money section):")
            print(body[:1500])
        return
    log_info(f"Extracted {len(bullets)} 'The Money' items")

    wb = None
    ws = None
    header_row = None
    if db_mode:
        headers = db_read_headers(db_path, database_url)
        hq_overrides = db_read_hq_overrides(db_path, database_url)
        db_hq_map = db_read_company_hq_map(db_path, database_url)
        db_sector_map = db_read_company_sector_map(db_path, database_url)
    else:
        log_info("Opening workbook")
        wb = load_workbook(args.path)
        if args.sheet not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{args.sheet}' not found in workbook")
        ws = wb[args.sheet]
        header_row = find_header_row(ws, "Company")
        headers = read_headers(ws, header_row)
        hq_overrides = {}
        db_hq_map = {}
        db_sector_map = {}

    hq_cache = load_hq_cache(args.hq_cache)
    log_info(f"HQ cache loaded entries: {len(hq_cache)}")
    if args.backfill_hq:
        if db_mode:
            raise RuntimeError("--backfill-hq is supported only in Excel mode")
        log_info("Running HQ backfill mode")
        updated = backfill_hq_from_cache(ws, header_row, headers, hq_cache)
        if args.dry_run:
            log_info(f"Dry run completed: would update {updated} HQ cells")
            print(f"Dry run: would update {updated} HQ cells")
            return
        wb.save(args.path)
        log_info(f"Workbook saved. Updated HQ cells: {updated}")
        print(f"Updated {updated} HQ cells")
        return

    log_info(f"Calling OpenAI model '{args.model}' to extract structured rows")
    rows = openai_extract_rows(bullets, headers, time_received, args.model)
    if not isinstance(rows, list):
        raise RuntimeError("OpenAI output is not a list")
    log_info(f"OpenAI returned {len(rows)} row candidates")

    for row in rows:
        # Force Tag empty
        if "Tag" in row:
            row["Tag"] = ""
        # Normalize Sector 1 to allowed list
        if "Sector 1" in row:
            sector = normalize_sector(row.get("Sector 1", ""))
            if not sector:
                sector = infer_sector_from_bullet(str(row.get("Company", "")), bullets)
            company_key = str(row.get("Company", "")).strip().lower()
            if company_key and company_key in db_sector_map:
                sector = db_sector_map[company_key]
            row["Sector 1"] = sector
        # Format Round size (€M) with comma decimal
        if "Round size (€M)" in row:
            row["Round size (€M)"] = format_decimal_comma(row.get("Round size (€M)", ""))
        # Fill HQ with priority: extracted value -> cache -> DB-known city -> bullet inference.
        company = str(row.get("Company", "")).strip()
        related_bullet = find_company_bullet(company, bullets)
        row["HQ"] = resolve_hq(company, row.get("HQ", ""), related_bullet, hq_overrides, hq_cache, db_hq_map)

    # Auto-enrich HQ city for unresolved companies and persist into hq_overrides.
    inferred_overrides = openai_enrich_hq_overrides(rows, bullets, args.hq_model)
    if inferred_overrides:
        log_info(f"Auto-enriched HQ city for {len(inferred_overrides)} companies")
        for row in rows:
            company = str(row.get("Company", "")).strip()
            if company in inferred_overrides:
                row["HQ"] = inferred_overrides[company]
        hq_overrides.update({k.strip().lower(): v for k, v in inferred_overrides.items()})
        if db_mode:
            db_upsert_hq_overrides(db_path, database_url, inferred_overrides)

    # Guardrail: skip malformed extracted rows (event-like noise without round size)
    rows = [
        r for r in rows
        if str(r.get("Company", "")).strip()
        and str(r.get("Round size (€M)", "")).strip()
    ]

    fallback_rows = synthesize_rows_for_missing_companies(
        bullets=bullets,
        rows=rows,
        headers=headers,
        date_value=parse_date_to_month_year(time_received),
        hq_overrides=hq_overrides,
        hq_cache=hq_cache,
        db_hq_map=db_hq_map,
    )
    if fallback_rows:
        log_info(f"Synthesized {len(fallback_rows)} fallback rows for missing companies")
        rows.extend(fallback_rows)

    # Final consistency pass: keep the canonical sector already used by the same company.
    for row in rows:
        company_key = str(row.get("Company", "")).strip().lower()
        if company_key and company_key in db_sector_map:
            row["Sector 1"] = db_sector_map[company_key]

    if db_mode:
        inserted, inserted_companies = db_insert_rows(
            db_path,
            headers,
            rows,
            dedup_company="Company",
            dedup_date="Date",
            database_url=database_url,
        )
    else:
        inserted, inserted_companies = append_rows(
            ws,
            header_row,
            headers,
            rows,
            dedup_company="Company",
            dedup_date="Date",
        )

    if args.dry_run:
        log_info(f"Dry run completed: would insert {inserted} rows")
        print(f"Dry run: would insert {inserted} rows")
        print("RESULT_JSON:" + json.dumps({"rows": inserted, "companies": inserted_companies}))
        return

    if db_mode:
        log_info(f"DB updated. Inserted rows: {inserted}. Companies: {', '.join(inserted_companies) if inserted_companies else '-'}")
    else:
        wb.save(args.path)
        log_info(f"Workbook saved. Inserted rows: {inserted}. Companies: {', '.join(inserted_companies) if inserted_companies else '-'}")
    print(f"Inserted {inserted} rows")
    print("RESULT_JSON:" + json.dumps({"rows": inserted, "companies": inserted_companies}))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
