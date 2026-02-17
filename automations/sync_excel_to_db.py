#!/usr/bin/env python3
import argparse
import os
import sqlite3
from typing import List
import difflib
import re

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

EXCEL_PATH_DEFAULT = "/Users/matteomoscarelli/Library/CloudStorage/OneDrive-Raccoltecondivise-UnitedVentures/United Ventures - United Ventures/04. Portfolio/Portfolio Team/100. OLD/07. Dealflow Meeting MM/Osservatorio VC Italy _.xlsx"
SHEET_DEFAULT = "Funding rounds (3)"
DB_PATH_DEFAULT = "/Users/matteomoscarelli/Documents/New project/db/rounds.db"

CANONICAL_SECTORS = [
    "Agritech",
    "Biotech",
    "Blue Economy",
    "Cleantech",
    "Climate Tech",
    "Consulting",
    "Consumer Services",
    "Crypto",
    "Cybersecurity",
    "Deep Tech",
    "DevOps",
    "eCommerce",
    "Edtech",
    "Energy",
    "Enterprise Tech",
    "Entraitment",
    "Fashion",
    "Femtech",
    "Fintech",
    "Food",
    "Gaming",
    "HR Tech",
    "Industrial Tech",
    "Insurtech",
    "Life Sciences",
    "Logistics",
    "Media & Ad",
    "Mobility",
    "Quantum",
    "Social Network",
    "Silver Economy",
    "Travel & Hospitality",
    "Web3",
]

_SECTOR_LOOKUP = {s.lower(): s for s in CANONICAL_SECTORS}

CITY_ALIASES = {
    "roma": "Rome",
    "rome": "Rome",
    "milano": "Milan",
    "milan": "Milan",
    "torino": "Turin",
    "turin": "Turin",
    "venezia": "Venice",
    "venice": "Venice",
    "firenze": "Florence",
    "florence": "Florence",
    "bologna": "Bologna",
    "parma": "Parma",
    "bergamo": "Bergamo",
    "como": "Como",
    "new york": "New York",
    "paris": "Paris",
}

MONTHS = {
    "jan": "Jan",
    "january": "Jan",
    "gen": "Jan",
    "gennaio": "Jan",
    "feb": "Feb",
    "february": "Feb",
    "febbraio": "Feb",
    "mar": "Mar",
    "march": "Mar",
    "marzo": "Mar",
    "apr": "Apr",
    "april": "Apr",
    "aprile": "Apr",
    "may": "May",
    "maggio": "May",
    "jun": "Jun",
    "june": "Jun",
    "giu": "Jun",
    "giugno": "Jun",
    "jul": "Jul",
    "july": "Jul",
    "lug": "Jul",
    "luglio": "Jul",
    "aug": "Aug",
    "august": "Aug",
    "ago": "Aug",
    "agosto": "Aug",
    "sep": "Sep",
    "september": "Sep",
    "set": "Sep",
    "settembre": "Sep",
    "oct": "Oct",
    "october": "Oct",
    "ott": "Oct",
    "ottobre": "Oct",
    "nov": "Nov",
    "november": "Nov",
    "novembre": "Nov",
    "dec": "Dec",
    "december": "Dec",
    "dic": "Dec",
    "dicembre": "Dec",
}


def normalize_sector(value: str) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    key = raw.lower()
    if key in _SECTOR_LOOKUP:
        return _SECTOR_LOOKUP[key]
    # Normalize common punctuation/spacing
    key = key.replace("-", " ").replace("_", " ").replace("&", "and")
    key = " ".join(key.split())
    if key in _SECTOR_LOOKUP:
        return _SECTOR_LOOKUP[key]
    # Fuzzy match to closest canonical sector
    candidates = difflib.get_close_matches(key, _SECTOR_LOOKUP.keys(), n=1, cutoff=0.85)
    if candidates:
        return _SECTOR_LOOKUP[candidates[0]]
    return raw


def normalize_key(value: str) -> str:
    if value is None:
        return ""
    raw = str(value)
    raw = re.sub(r"\s+", " ", raw).strip()
    raw = raw.replace("\u200b", "")
    raw = re.sub(r"[^\w\s&.+-]", "", raw, flags=re.UNICODE)
    return raw.casefold()


def normalize_city(value: str) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    if "," in raw:
        return raw
    key = normalize_key(raw)
    return CITY_ALIASES.get(key, raw)


def normalize_date(value: str) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    # If already in "Mon YYYY"
    m = re.search(r"\b([A-Za-z]{3})\s+(20\d{2})\b", raw)
    if m:
        mon = MONTHS.get(m.group(1).lower(), m.group(1).title())
        return f"{mon} {m.group(2)}"
    # YYYY-MM-DD or YYYY/MM/DD
    m = re.search(r"\b(20\d{2})[-/](\d{2})[-/]\d{2}\b", raw)
    if m:
        year, month = m.group(1), m.group(2)
        mon = MONTHS.get(month, None)
        if mon is None:
            mon = list(MONTHS.values())[int(month) - 1]
        return f"{mon} {year}"
    # Month name + year
    m = re.search(r"\b([A-Za-zÀ-ÿ]+)\s+(20\d{2})\b", raw)
    if m:
        mon = MONTHS.get(m.group(1).lower(), m.group(1).title())
        return f"{mon} {m.group(2)}"
    return raw


def normalize_round_size(value: str) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    # Keep only numeric forms and normalize decimal separator to dot for DB consistency.
    cleaned = raw.replace(",", ".")
    if re.fullmatch(r"\d+(\.\d+)?", cleaned):
        return cleaned
    return raw


def find_header_row(ws, header_name: str) -> int:
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == header_name.lower():
                return cell.row
    raise RuntimeError(f"Header '{header_name}' not found")


def read_headers(ws, header_row: int) -> List[str]:
    headers = []
    for cell in ws[header_row]:
        if cell.value is None:
            headers.append("")
        else:
            headers.append(str(cell.value).strip())
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def sync_excel_to_db(excel_path: str, sheet: str, db_path: str) -> int:
    if load_workbook is None:
        raise RuntimeError("openpyxl not installed")

    wb = load_workbook(excel_path)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet}' not found in workbook")
    ws = wb[sheet]

    header_row = find_header_row(ws, "Company")
    headers = read_headers(ws, header_row)
    if not headers:
        raise RuntimeError("No headers found")

    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Create table with quoted column names
    cols_sql = ", ".join([f'"{h}" TEXT' for h in headers])
    cur.execute("DROP TABLE IF EXISTS rounds")
    cur.execute(f"CREATE TABLE rounds (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols_sql})")

    inserted = 0
    canonical_maps = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        values = []
        empty = True
        for idx, h in enumerate(headers):
            cell = row[idx].value if idx < len(row) else None
            if cell not in (None, ""):
                empty = False
            header = h.strip()
            header_lc = header.lower()
            if header_lc == "sector 1":
                values.append(normalize_sector(cell))
                continue
            if header_lc == "hq":
                values.append(normalize_city(cell))
                continue
            if header_lc == "date":
                values.append(normalize_date(cell))
                continue
            if header_lc == "round size (€m)":
                values.append(normalize_round_size(cell))
                continue
            if cell is None:
                values.append("")
                continue
            raw = str(cell).strip()
            if header_lc in {"company", "lead", "co-lead / follow 1", "follow 2", "follow 3", "follow 4", "debt"}:
                cmap = canonical_maps.setdefault(header, {})
                key = normalize_key(raw)
                if key in cmap:
                    values.append(cmap[key])
                else:
                    cmap[key] = raw
                    values.append(raw)
            else:
                values.append(raw)
        if empty:
            continue
        placeholders = ", ".join(["?"] * len(headers))
        cur.execute(f"INSERT INTO rounds ({', '.join([f'"{h}"' for h in headers])}) VALUES ({placeholders})", values)
        inserted += 1

    conn.commit()
    conn.close()
    return inserted


def main():
    parser = argparse.ArgumentParser(description="Sync Excel to SQLite DB.")
    parser.add_argument("--path", default=EXCEL_PATH_DEFAULT)
    parser.add_argument("--sheet", default=SHEET_DEFAULT)
    parser.add_argument("--db", default=DB_PATH_DEFAULT)
    args = parser.parse_args()

    inserted = sync_excel_to_db(args.path, args.sheet, args.db)
    print(f"Synced {inserted} rows to DB")


if __name__ == "__main__":
    main()
